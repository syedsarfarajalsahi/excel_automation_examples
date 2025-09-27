import pandas as pd
from openpyxl import Workbook
from datetime import datetime
import os


# 1. Load the orders CSV
orders = pd.read_csv("orders.csv")


# 2. Create output folder if not exists
os.makedirs("invoices", exist_ok=True)


# 3. Group orders by customer name
grouped_orders = orders.groupby("customer_name")


# 4. Generate invoice for each customer
for customer, data in grouped_orders:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Add header info
    ws["A1"] = "INVOICE"
    ws["A2"] = f"Customer: {customer}"
    ws["A3"] = f"Date: {datetime.today().strftime('%Y-%m-%d')}"

    # Table headers
    ws.append(["Product", "Quantity", "Price", "Total"])

    total_amount = 0

    # Add each product row
    for _, row in data.iterrows():
        product = row["product"]
        quantity = row["quantity"]
        price = row["price"]
        total = quantity * price
        total_amount += total
        ws.append([product, quantity, price, total])

    # Empty row before grand total
    ws.append([])

    # Add grand total row
    ws.append(["", "", "Grand Total:", total_amount])

    # Make filename safe
    safe_name = customer.replace(" ", "_").replace(".", "")
    filename = f"invoices/invoice_{safe_name}.xlsx"

    # Save invoice
    wb.save(filename)
    print(f"Invoice generated for {customer}: {filename}")

# 5. Confirmation message
print("\nAll invoices generated successfully!")
